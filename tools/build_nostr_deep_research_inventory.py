from __future__ import annotations

import html
import json
import re
import socket
import ssl
import time
from collections import Counter, defaultdict, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from html.parser import HTMLParser
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qsl, urlencode, urldefrag, urljoin, urlparse, urlunparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(r"C:\Users\34669\OneDrive - Crays Europe SE\Desktop\nostr_deep_research_linkdatenbank.xlsx")
OUT = ROOT / "tools" / "nostr_deep_research_inventory.json"
USER_AGENT = "CraysNostrDeepResearch/1.0 (+https://www.crays.org/nostr/)"
URL_RE = re.compile(r"https?://[^\s,;\]>)\"]+")

SUBCRAWL_DOMAINS = {
    "nostr.com",
    "www.nostr.com",
    "nostr.org",
    "www.nostr.org",
    "nostr.how",
    "www.nostr.how",
    "nostr.net",
    "www.nostr.net",
    "start.nostr.net",
    "www.start.nostr.net",
    "nostrapps.com",
    "www.nostrapps.com",
    "nostrlogin.org",
    "www.nostrlogin.org",
    "nostr.co.uk",
    "www.nostr.co.uk",
    "nostr.watch",
    "www.nostr.watch",
    "nostr.directory",
    "www.nostr.directory",
    "nostrcompass.org",
    "www.nostrcompass.org",
    "nostrdesign.org",
    "www.nostrdesign.org",
    "hellonostr.dev",
    "www.hellonostr.dev",
    "primal.net",
    "www.primal.net",
}

GLOBAL_SUBPAGE_LIMIT = 5200
PER_DOMAIN_SUBPAGE_LIMIT = 260
PER_SOURCE_SUBPAGE_LIMIT = 90

SKIP_EXTENSIONS = {
    ".7z",
    ".avi",
    ".bmp",
    ".css",
    ".dmg",
    ".exe",
    ".gif",
    ".gz",
    ".ico",
    ".jpeg",
    ".jpg",
    ".js",
    ".json",
    ".mov",
    ".mp3",
    ".mp4",
    ".pdf",
    ".png",
    ".rss",
    ".svg",
    ".tar",
    ".tgz",
    ".webm",
    ".webp",
    ".zip",
}

SKIP_PATH_TOKENS = {
    "/actions",
    "/commits",
    "/compare",
    "/forks",
    "/graphs",
    "/issues",
    "/network",
    "/pulls",
    "/pulse",
    "/releases",
    "/security",
    "/stargazers",
    "/watchers",
    "/login",
    "/signup",
    "/search",
    "/cart",
    "/checkout",
}


def clean_text(value: object) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slugify(value: str) -> str:
    value = clean_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return value or "item"


def normalize_url(url: str) -> str:
    url, _fragment = urldefrag((url or "").strip())
    parsed = urlparse(url)
    if not parsed.scheme:
        parsed = urlparse(f"https://{url}")
    scheme = parsed.scheme.lower()
    netloc = parsed.netloc.lower()
    path = parsed.path or "/"
    if path != "/" and path.endswith("/"):
        path = path.rstrip("/")
    query_pairs = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if not key.lower().startswith("utm_") and key.lower() not in {"ref", "fbclid", "gclid"}
    ]
    query = urlencode(query_pairs, doseq=True)
    return urlunparse((scheme, netloc, path, "", query, ""))


def same_site_link(base: str, href: str) -> str | None:
    if not href or href.startswith(("mailto:", "tel:", "javascript:", "nostr:", "data:")):
        return None
    url = normalize_url(urljoin(base, href))
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return None
    if any(parsed.path.lower().endswith(ext) for ext in SKIP_EXTENSIONS):
        return None
    return url


class ResearchPageParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.title = ""
        self.description = ""
        self.headings: list[dict[str, str]] = []
        self.links: list[str] = []
        self.paragraphs: list[str] = []
        self._capture: str | None = None
        self._heading_level = ""
        self._buffer: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {name.lower(): value or "" for name, value in attrs}
        if tag in {"script", "style", "noscript", "svg"}:
            self._skip_depth += 1
            return
        if self._skip_depth:
            return
        if tag == "a" and attrs_dict.get("href"):
            self.links.append(attrs_dict["href"])
        if tag == "meta":
            name = attrs_dict.get("name", "").lower()
            prop = attrs_dict.get("property", "").lower()
            if name == "description" or prop == "og:description":
                if not self.description:
                    self.description = clean_text(attrs_dict.get("content", ""))
        if tag == "title":
            self._capture = "title"
            self._buffer = []
        elif tag in {"h1", "h2", "h3"}:
            self._capture = "heading"
            self._heading_level = tag
            self._buffer = []
        elif tag == "p":
            self._capture = "paragraph"
            self._buffer = []

    def handle_endtag(self, tag: str) -> None:
        if self._skip_depth:
            if tag in {"script", "style", "noscript", "svg"}:
                self._skip_depth -= 1
            return
        if self._capture == "title" and tag == "title":
            self.title = clean_text(" ".join(self._buffer))
            self._capture = None
        elif self._capture == "heading" and tag == self._heading_level:
            text = clean_text(" ".join(self._buffer))
            if text:
                self.headings.append({"level": self._heading_level, "text": text})
            self._capture = None
            self._heading_level = ""
        elif self._capture == "paragraph" and tag == "p":
            text = clean_text(" ".join(self._buffer))
            if len(text) > 60:
                self.paragraphs.append(text)
            self._capture = None

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        if self._capture:
            self._buffer.append(data)


def fetch_page(url: str, timeout: int = 18) -> dict:
    req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml,text/plain;q=0.9,*/*;q=0.2"})
    started = time.perf_counter()
    try:
        with urlopen(req, timeout=timeout, context=ssl.create_default_context()) as response:
            final_url = response.geturl()
            content_type = response.headers.get("content-type", "")
            raw = response.read(1_500_000)
            status = getattr(response, "status", 200)
    except (HTTPError, URLError, TimeoutError, OSError, socket.timeout) as exc:
        return {
            "url": url,
            "status": "error",
            "http_status": getattr(exc, "code", None),
            "error": clean_text(str(exc))[:320],
            "elapsed_ms": int((time.perf_counter() - started) * 1000),
        }
    text = raw.decode("utf-8", "replace")
    parser = ResearchPageParser()
    if "html" in content_type.lower() or text.lstrip().startswith(("<!doctype", "<html", "<")):
        try:
            parser.feed(text)
        except Exception:
            pass
    title = parser.title or final_url
    description = parser.description
    if not description and parser.paragraphs:
        description = parser.paragraphs[0][:280]
    links = []
    for href in parser.links:
        normalized = same_site_link(final_url, href)
        if normalized:
            links.append(normalized)
    return {
        "url": url,
        "final_url": final_url,
        "status": "ok",
        "http_status": status,
        "content_type": content_type,
        "title": clean_text(title)[:220],
        "description": clean_text(description)[:360],
        "headings": parser.headings[:32],
        "paragraph_samples": [clean_text(p)[:320] for p in parser.paragraphs[:4]],
        "links": sorted(set(links))[:220],
        "html_bytes": len(raw),
        "elapsed_ms": int((time.perf_counter() - started) * 1000),
    }


def workbook_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    row_id = 0
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        headers = [clean_text(ws.cell(1, c).value) or f"Column {c}" for c in range(1, ws.max_column + 1)]
        for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            urls: list[str] = []
            for value in values:
                if isinstance(value, str):
                    urls.extend(URL_RE.findall(value))
            if not urls:
                continue
            row_id += 1
            rows.append(
                {
                    "row_id": row_id,
                    "sheet": ws.title,
                    "excel_row": excel_row,
                    "category": clean_text(data.get("Category") or data.get("Topic cluster") or data.get("NIP") or ws.title),
                    "subcategory": clean_text(data.get("Subcategory") or data.get("Status") or ""),
                    "name": clean_text(data.get("Name") or data.get("Title") or data.get("NIP") or ""),
                    "importance": clean_text(data.get("Importance") or data.get("Status") or ""),
                    "notes": clean_text(data.get("Notes") or ""),
                    "source_evidence": clean_text(data.get("Source / evidence") or ""),
                    "urls": [normalize_url(url) for url in urls],
                    "raw": {key: clean_text(value) for key, value in data.items() if clean_text(value)},
                }
            )
    return rows


def importance_rank(value: str) -> int:
    value = (value or "").lower()
    if "core" in value:
        return 0
    if "high" in value or "active" in value:
        return 1
    if "medium" in value:
        return 2
    return 3


def build_sources(rows: list[dict]) -> dict[str, dict]:
    grouped: dict[str, dict] = {}
    for row in rows:
        for position, url in enumerate(row["urls"]):
            source = grouped.setdefault(
                url,
                {
                    "url": url,
                    "primary_names": [],
                    "categories": [],
                    "subcategories": [],
                    "importance_values": [],
                    "notes": [],
                    "source_evidence": [],
                    "row_refs": [],
                    "secondary_urls": [],
                },
            )
            if row["name"]:
                source["primary_names"].append(row["name"])
            if row["category"]:
                source["categories"].append(row["category"])
            if row["subcategory"]:
                source["subcategories"].append(row["subcategory"])
            if row["importance"]:
                source["importance_values"].append(row["importance"])
            if row["notes"]:
                source["notes"].append(row["notes"])
            if row["source_evidence"]:
                source["source_evidence"].append(row["source_evidence"])
            source["row_refs"].append({"sheet": row["sheet"], "row": row["excel_row"], "name": row["name"], "category": row["category"]})
            for other in row["urls"]:
                if other != url:
                    source["secondary_urls"].append(other)
            if position == 0:
                source["primary_workbook_url"] = True
    for source in grouped.values():
        for key in ["primary_names", "categories", "subcategories", "importance_values", "notes", "source_evidence", "secondary_urls"]:
            source[key] = sorted(set(filter(None, source[key])))
        source["importance_rank"] = min([importance_rank(value) for value in source["importance_values"]] or [3])
    return grouped


def should_subcrawl(url: str, source: dict) -> bool:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    if any(parsed.path.lower().endswith(ext) for ext in SKIP_EXTENSIONS):
        return False
    return True


def github_repo_scope(url: str) -> str | None:
    parsed = urlparse(url)
    if parsed.netloc not in {"github.com", "www.github.com"}:
        return None
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) < 2:
        return None
    return "/" + "/".join(parts[:2])


def link_in_source_scope(parent_url: str, link: str) -> bool:
    parent = urlparse(parent_url)
    child = urlparse(link)
    if child.scheme not in {"http", "https"}:
        return False
    path = child.path.lower()
    if any(path.endswith(ext) for ext in SKIP_EXTENSIONS):
        return False
    if any(token in path for token in SKIP_PATH_TOKENS):
        return False
    if child.netloc != parent.netloc:
        return False
    repo_scope = github_repo_scope(parent_url)
    if repo_scope:
        return child.path == repo_scope or child.path.startswith(repo_scope + "/")
    return True


def fetch_many(urls: list[str], max_workers: int = 10) -> dict[str, dict]:
    results: dict[str, dict] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_url = {pool.submit(fetch_page, url): url for url in urls}
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                results[url] = future.result()
            except Exception as exc:
                results[url] = {"url": url, "status": "error", "error": clean_text(str(exc))[:320]}
    return results


def crawl_subpages(sources: dict[str, dict], fetched: dict[str, dict]) -> dict[str, list[dict]]:
    per_source: dict[str, list[dict]] = defaultdict(list)
    per_domain_count: Counter[str] = Counter()
    per_source_count: Counter[str] = Counter()
    queued: deque[tuple[str, str, int]] = deque()
    seen: set[str] = set(fetched)
    for url, source in sources.items():
        if not should_subcrawl(url, source):
            continue
        record = fetched.get(url, {})
        for link in record.get("links", [])[:180]:
            parsed = urlparse(link)
            if link_in_source_scope(url, link) and link not in seen:
                queued.append((url, link, 1))
                seen.add(link)
    while queued and len(seen) < GLOBAL_SUBPAGE_LIMIT:
        parent, url, depth = queued.popleft()
        domain = urlparse(url).netloc
        if per_domain_count[domain] >= PER_DOMAIN_SUBPAGE_LIMIT:
            continue
        if per_source_count[parent] >= PER_SOURCE_SUBPAGE_LIMIT:
            continue
        record = fetch_page(url, timeout=14)
        per_domain_count[domain] += 1
        per_source_count[parent] += 1
        summary = {
            "url": url,
            "status": record.get("status"),
            "http_status": record.get("http_status"),
            "title": record.get("title", url),
            "description": record.get("description", ""),
            "headings": record.get("headings", [])[:8],
        }
        per_source[parent].append(summary)
        if depth < 2 and record.get("status") == "ok":
            for link in record.get("links", [])[:90]:
                if link in seen:
                    continue
                if link_in_source_scope(parent, link):
                    queued.append((parent, link, depth + 1))
                    seen.add(link)
        time.sleep(0.02)
    return per_source


def main() -> None:
    workbook_path = DEFAULT_WORKBOOK
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    rows = workbook_rows(workbook_path)
    sources = build_sources(rows)
    urls = sorted(sources, key=lambda url: (sources[url].get("importance_rank", 3), url))
    fetched = fetch_many(urls)
    subpages = crawl_subpages(sources, fetched)
    for url, source in sources.items():
        source["fetch"] = fetched.get(url, {"status": "not checked"})
        source["subpages_checked"] = subpages.get(url, [])
    inventory = {
        "generated_at": "2026-05-31",
        "workbook": str(workbook_path),
        "workbook_rows_with_urls": len(rows),
        "unique_urls": len(sources),
        "url_cells": sum(len(row["urls"]) for row in rows),
        "sources": sorted(sources.values(), key=lambda item: (item.get("importance_rank", 3), item.get("categories", [""])[0], item.get("primary_names", [""])[0], item["url"])),
        "summary": {
            "by_category": Counter(category for source in sources.values() for category in source["categories"]),
            "by_importance": Counter(value for source in sources.values() for value in source["importance_values"]),
            "fetch_status": Counter(source["fetch"].get("status", "unknown") for source in sources.values()),
            "subpages_checked": sum(len(source.get("subpages_checked", [])) for source in sources.values()),
        },
    }
    OUT.write_text(json.dumps(inventory, indent=2, ensure_ascii=False), encoding="utf-8")
    print(
        "Built deep research inventory:",
        inventory["workbook_rows_with_urls"],
        "rows,",
        inventory["unique_urls"],
        "unique URLs,",
        inventory["summary"]["fetch_status"],
        "fetches,",
        inventory["summary"]["subpages_checked"],
        "subpages checked",
    )


if __name__ == "__main__":
    main()
